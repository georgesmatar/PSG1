#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Fri May 26 12:09:20 2017

@author: Tomy Aumont
"""
import os
import numpy as np
import pandas
import copy as cp
import datetime as dt
import openpyxl
import matplotlib.pyplot as plt
from openpyxl.utils.dataframe import dataframe_to_rows

plt.rcParams.update({'figure.max_open_warning': 0})

def CreateFolderArchitecture(FS_method):
    if not os.path.exists("Results"):
        os.mkdir("Results")
    currentDateTime = str(dt.datetime.now().strftime("%d%b%Y_%H-%M-%S"))
    resultDir = "Results/" + currentDateTime + '_' + FS_method + '/'
    os.mkdir(resultDir)

    return currentDateTime,resultDir

def RemoveNaNs(X,method,debug_flag):
    if method == 'Feature':
        rmTag = 'Removed_Features'
        removedData = X.columns[X.isnull().any()].tolist()
        X.dropna(axis=1,how='any',inplace=True)
    elif method == 'Subject': # error : not removing enough subjects...
        rmTag = 'Removed_Subject'
        removedData = X.index[X.isnull().any()].tolist()
        X.dropna(axis=0,how='any',inplace=True)

    print('Removed items : {}'.format(removedData))

    # Get the tag of the removed items, either features or subjects
    tmp=[]
    for i in list(range(len(removedData))):
        tmp.append(removedData[i])
    removedInfo = pandas.DataFrame(data=tmp,columns=[rmTag])

    return X, removedInfo

def GetClasses(dataset, classID):
    if classID == 'Sexe' or classID == 'Groupe' :
        tmp=cp.deepcopy(dataset[classID])
        y = pandas.DataFrame(data=np.reshape(tmp,[dataset.shape[0],1]),
                             columns=['Cluster'])
    elif classID == 'Mix' :
        tmp = []
        for i in list(range(len(dataset))):
            if dataset['Groupe'].iloc[i]==1 and dataset['Sexe'].iloc[i]==1:
                tmp.append([0,'youngWoman'])
            elif dataset['Groupe'].iloc[i]==2 and dataset['Sexe'].iloc[i]==1:
                tmp.append([1,'oldWoman'])
            elif dataset['Groupe'].iloc[i]==1 and dataset['Sexe'].iloc[i]==2:
                tmp.append([2,'youngMan'])
            elif dataset['Groupe'].iloc[i]==2 and dataset['Sexe'].iloc[i]==2:
                tmp.append([3,'oldMan'])
        y = pandas.DataFrame(data=tmp,columns=['Cluster','Cluster_Name'])

    return y

def GetSamples(dataset):
    X = cp.deepcopy(dataset)
    # Remove class ID columns
    X.drop(['Groupe','Sexe'],1, inplace=True)

    return X
"""
   Balance class so there is as many old men and old women in dataset
"""
def BalanceClasses(x,y):
    import random as rd

    X = cp.deepcopy(x)
    Y = cp.deepcopy(y)
    #    oldMen = y[y['Cluster_Name']=='oldMan']
    #    oldWomen = y[y['Cluster_Name']=='oldWoman']
    d = len(Y[Y['Cluster_Name']=='oldWoman']) - \
    len(Y[Y['Cluster_Name']=='oldMan'])
    if d > 0:
        idx = rd.sample(set(Y[Y['Cluster_Name']=='oldWoman'].index),d)
    else:
        idx = rd.sample(set(Y[Y['Cluster_Name']=='oldMan'].index),np.abs(d))
    # Ne veux pas dropper correctement...
    print('Dropped subject indexes: {}'.format(idx))
    X.drop(X.index[idx],inplace=True)
    Y.drop(idx,inplace=True)

    return X, Y

def Standardize(dataset,classes,method,debug_flag):
#    from scipy.stats import zscore
    # Not adapted for 4 groups classification
    if method == 'Across_Group': # Standardize by group
#        print('Standardize data across group of subjects')
#        uniqueY = np.unique(classes)
#        xSet = pandas.DataFrame()
#        for i in uniqueY:
#            tmp = pandas.DataFrame(data=cp.deepcopy(dataset[list(classes==i)]),
#                                   columns = ['Classe_'+str(i)])
#            print(tmp)
#            xSet.append(tmp)
##            xSet  =[xSet, newSet]
#            print(xSet)
        # Copy data of the two classes separatly
        youngX = cp.deepcopy(dataset[list(classes==1)]) # class 1 data
        oldX   = cp.deepcopy(dataset[classes==2]) # class 2 data

        # Remove class ID columns
        youngX.drop(['Groupe','Sexe'],1, inplace=True)
        oldX.drop(['Groupe','Sexe'],1, inplace=True)
        if debug_flag:
            print("Young without 'Group' and 'Sexe' column")
            print(youngX)
            print("Old without 'Group' and 'Sexe' column")
            print(oldX)
            print("Classes")

        # separate standardization with z-score
        youngZX,zParm1 = ComputeZScore(youngX)
        oldZX, zParam2 = ComputeZScore(oldX)
        zParam = [zParm1, zParam2]
        if debug_flag:
            print("Young standardized without 'Group' and 'Sexe' column")
            print(youngZX)
            print("Old standardized without 'Group' and 'Sexe' column")
            print(oldZX)

        # Combine young and old into a single matrix
        X = youngZX.append(oldZX)
        if debug_flag:
            print("Combined standardized data without 'Group' and 'Sexe' column")
            print(youngZX)

    elif method == 'Across_Subjects': # Standardization across all subject
#        print('Standardization data across all subjects')
        # Copy all dataset
        X = cp.deepcopy(dataset)
        # Remove class ID columns
#        X.drop(['Groupe','Sexe'],1, inplace=True)
        if debug_flag:
            print("All data without 'Group' and 'Sexe' column")
            print(X)

        # group standardization with z-score
        X, zParam = ComputeZScore(X, debug_flag)

        if debug_flag:
            print("All standardized data without 'Group' and 'Sexe' column")
            print(X)
            print('Standardization parameters')
            print(zParam)
    else: # not standardized
#        print('No standardization')
        X = cp.deepcopy(dataset)
        # Remove class ID columns
        X.drop(['Groupe','Sexe'],1, inplace=True)

        zParam = []

    return X, zParam

"""
Compute z-score on features that have unit

Return :    z-score transformation of dataset
            z-score parameter per feature
"""
def ComputeZScore(dataset, debug_flag=0):

    mns = []
    sstd = []
    result = cp.deepcopy(dataset)

    for i in list(dataset):
        if (i != 'EffSommeil' and i != 'Qualite subj.sommeil' and
            i != 'PSQI' and i != 'Rapport delta-theta' and
            i != 'Rapport theta-alpha' and i != 'Rapport theta-sigma' and
            i != 'Rapport theta-beta' and i != 'Rapport alpha-sigma' and
            i != 'Rapport alpha-beta' and i != 'Rapport sigma-beta'):
            m = dataset[i].mean()
            s = dataset[i].std()
            result[i] = (dataset[i] - m) / s
            mns.append(m)
            sstd.append(s)
        else:
            mns.append('')
            sstd.append('')

    if debug_flag:
        print('mean = {}'.format(mns))
        print('end mean')
        print('std = {}'.format(sstd))
        print('end std')

    tmp1 = pandas.DataFrame(data=mns,columns=['mean'])
    tmp2 = pandas.DataFrame(data=sstd,columns=['std'])
    param = pandas.concat([tmp1,tmp2],axis=1)
    return result, param

"""
Apply z-score stanrdadization to 'a' using mean and std from 'p'
"""
def ApplyStandardization(a,p):
    res = cp.deepcopy(a)
    for s in list(range(a.shape[0])):
        for f in list(range(a.shape[1])):
            if (p['mean'][f] != '' and p['std'][f] != ''):
                res.iloc[s][f] = (a.iloc[s][f]-p['mean'][f])/p['std'][f]
    return res

def ComputeRatio(X):
    X['Rapport delta-theta'] = X['Delta abs'] / X['Thêta abs']
    X['Rapport theta-alpha'] = X['Thêta abs'] / X['Alpha abs']
    X['Rapport theta-sigma'] = X['Thêta abs'] / X['Sigma abs']
    X['Rapport theta-beta']  = X['Thêta abs'] / X['Bêta abs']
    X['Rapport alpha-sigma'] = X['Alpha abs'] / X['Sigma abs']
    X['Rapport alpha-beta']  = X['Alpha abs'] / X['Bêta abs']
    X['Rapport sigma-beta']  = X['Sigma abs'] / X['Bêta abs']
    return X

def Permute(clusters,xTrain,xTest,yTrain,yTest,N,standardization,debug_flag=0):

    DA = [] # permutation's decoding accuracy

#    # Create subset based on selected best features
#    xTrain_estim = estimator.transform(xTrain)
#    xTest_estim = estimator.transform(xTest)
    # Create a classifier trainned with permutted label
    # Create the estimator and RFE object with a cross-validated score.
    if clusters == 'Mix':
        from sklearn.svm import LinearSVC
        permClf = LinearSVC(dual=False,multi_class='ovr')
    else:
        from sklearn.svm import SVC
        permClf = SVC(kernel="linear",shrinking=False)

    print('Permutting')
    for permIt in list(range(1,N+1)):
        print('\rPermutation {} of {} \r'.format(permIt,N),flush=True)

        # randomly permutte label to create the 'luck' probability
        permutted_yTrain = np.random.permutation(yTrain)
        # Data z-score standardization
        xTrainSet,zPrm = Standardize(xTrain,permutted_yTrain,standardization,0)
        # train classifier with permutted label
        permClf = permClf.fit(xTrainSet,permutted_yTrain)


        # standardize test set using trainset standardization parameters
        xTestSet = ApplyStandardization(xTest,zPrm)
#        # Generate the new subsets based on the selected features
#        X_train_sfs = permClf.transform(xTrainSet.as_matrix())
#        X_test_sfs = permClf.transform(xTestSet.as_matrix())

#        # Fit the estimator using the new feature subset
#        # and make a prediction on the test data
#        permClf.fit(X_train_sfs, permutted_yTrain)
        y_pred = permClf.predict(xTestSet)

        # Compute the accuracy of the test prediction
        acc = float((yTest == y_pred).sum()) / y_pred.shape[0]
        if debug_flag:
            print('\nPSet accuracy\t: %.2f %%' % (acc * 100), end='\r\r')
        DA.append(acc)


#        # test classifier
#        y_pred = permClf.predict(xTest)
#        # Compute the accuracy of the prediction
#        acc = float((yTest == y_pred).sum()) / y_pred.shape[0]
#        if debug_flag:
#            print('Permutation #%d set accuracy\t: %.2f %%' % \
#                  (permIt,(acc * 100)))
#        DA.append(acc)
    print('')
    return DA

def ComputePermutationAvgDA(avgDA):
    # Create a dataframe with the received avergae DA
    DA1=pandas.DataFrame(data=avgDA,columns=['Avg_Permutation_DA_per_epoch'])
    # Add column containing the computed average DA of all iteration
    DA2 = pandas.DataFrame(data=[np.mean(DA1['Avg_Permutation_DA_per_epoch'])],
                           columns=['Global_Permutation_DA'])
    DA = pandas.concat([DA1,DA2],axis=1)

    return DA

def SaveExcel(fileName,sheetName,data,debug_flag=0):
    print('\nWritting the excel file')
    # load existing excel workbook or create it
    if os.path.isfile(fileName):
        workbook = openpyxl.load_workbook(fileName,0)
    else:
        workbook = openpyxl.Workbook()

    # create a new worksheet
    workbook.create_sheet(sheetName)
    # convert dataframe into something suitable for an excel worksheet
    rows = dataframe_to_rows(data,index=False)

    # Write new results to excel
    if debug_flag:
        print('DAs to write :')

    for r_idx, row in enumerate(rows, 1):
        for c_idx, value in enumerate(row, 1):
            # write scores and ranks
             workbook[sheetName].cell(row=r_idx, column=c_idx, value=value)

    # save excel with modifications
    workbook.save(fileName)

def PlotBestCombinationMetrics(data,figName):
    # Plot mean/std
    plt.figure(dpi=300)
    plt.title('Metrique choix meilleure combinaison')
    plt.xlabel("nb attributs dans combinaison")
    plt.xticks(range(data.shape[0]))
    plt.ylabel("Moyenne sur ecart-type")
    for n in list(range(data.shape[1])):
        plt.plot(list(range(1,data.shape[0]+1)),data.iloc[:,n])
    plt.savefig(figName, bbox_inches='tight')
    plt.clf()
    plt.close()

def PlotPermHist(data,testAvgDA,dateTime,savedPlotName):
    from math import floor

    if len(np.shape(data)) == 2:
        df = pandas.DataFrame(data=data)
        dataHist, bins = np.histogram(df,bins=100,range=(0,1))
        pValue = dataHist[floor(testAvgDA*100):].sum()/dataHist.sum()

    elif len(np.shape(data)) > 2: # sum histograms bins for each epoch
        df = pandas.DataFrame(data=data)
        df = df.T
        dataHist = pandas.DataFrame()
        for i in list(range(len(df))):
            tmp, bins = np.histogram(df.iloc[i][:],bins=100,range=(0,1))
            dataHist = tmp if i == 0 else dataHist + tmp

        pValue = dataHist[floor(testAvgDA.iloc[0][1]*100),:].sum() / \
                    dataHist.sum()

    print('\nPermutation pValue : {}'.format(pValue))

    plt.figure(figsize=[10,8],dpi=300)
    plt.title('Permutation decoding accuracy\nN={} pValue = {}\n {}'.format(
            dataHist.sum(),pValue,dateTime))
    plt.xlabel("Decoding accuracy")
    plt.ylabel("Number of occurence")
    plt.bar(range(100),dataHist,width=0.9,label='Occurances',align='center')
    plt.axvline(x=testAvgDA*100,color='r',label='test DA')
    plt.legend()
    plt.tight_layout()
    plt.savefig(savedPlotName, bbox_inches='tight')
    plt.clf()
    print('\tComplete')
    return plt


def PlotHist(dat,xLabels,title,savedPlotName):#,maxY

    manualFlag = 0
    # for manual use
    if manualFlag:
        dat = pandas.read_excel('filename.xlsx',header=0)
        occ = dat['Occurence_Best']
        names = dat['Features_Name']
        data = occ.dropna()
    else:
        data = cp.deepcopy(dat).dropna()
        labels = pandas.DataFrame(data=xLabels)
    
    data.sort_values(ascending=False,inplace=True)

    if manualFlag:
        xTickLabel = names[data.index]
    else:
        tmp = labels.iloc[data.index]
        xTickLabel = tmp

    plt.figure(dpi=300)
    plt.title(title)
    plt.xlabel("Features")
    plt.xticks(range(len(data)),xTickLabel,rotation='vertical')
    plt.ylabel("Occurence")
    plt.bar(range(len(data)),data,width=0.9,label='Occurences',align='center')
    plt.legend()
    plt.savefig(savedPlotName, bbox_inches='tight')
    plt.clf()
    plt.close()

def plot_confusion_matrix(cm, classes,
                          normalize=False,
                          precision=8,
                          title='Confusion matrix',
                          cmap=plt.cm.Blues):
    """
    This function prints and plots the confusion matrix.
    Normalization can be applied by setting `normalize=True`.
    """
    import itertools

    np.set_printoptions(precision=2)

    if normalize:
        cm = cm.astype('float') / cm.sum(axis=1)[:, np.newaxis]
        print("Normalized confusion matrix")
    else:
        print('Confusion matrix, without normalization')

    plt.imshow(cm, interpolation='nearest', cmap=cmap)
    plt.title(title)
    plt.colorbar()
    tick_marks = np.arange(len(classes))
    plt.xticks(tick_marks, classes, rotation=45)
    plt.yticks(tick_marks, classes)

    print(cm)

    thresh = cm.max() / 2.
    for i, j in itertools.product(range(cm.shape[0]), range(cm.shape[1])):
        plt.text(j, i, round(cm[i, j],precision),
                 horizontalalignment="center",
                 color="white" if cm[i, j] > thresh else "black")

    plt.tight_layout()
    plt.ylabel('True label')
    plt.xlabel('Predicted label')
    plt.savefig(title+'.png', bbox_inches='tight')
    plt.clf()
    plt.close()

    return plt

def GetSubsetOccurence(bestSets):
    """
    Ne depasse pas les combinaisons de 8 attributs, a 9 la memoire bust
    """
    import itertools
#    import numpy as np
    import pandas as pd
    from gc import collect as clr


#    bestSets = [(1,2,3),(2,4,3),(3,4,5),(4,6,5),(5,6,7)]

    #**************** only to test if not running whole script ****************
    # 10 iterations
#    bestSets = [(2, 4, 11, 14, 15, 17, 18, 19, 21, 22, 23, 24, 26, 27, 28, 29, 38, 39),(0, 2, 3, 4, 5, 6, 9, 10, 12, 13, 14, 18, 19, 21, 22, 23, 24, 26, 27, 28, 30, 36, 37, 38),(1, 2, 4, 8, 9, 10, 11, 13, 14, 15, 16, 19, 20, 21, 22, 24, 26, 27, 28, 30, 33, 34, 35, 38, 39),(1, 2, 3, 4, 5, 8, 11, 13, 14, 15, 18, 19, 20, 21, 22, 23, 24, 25, 26, 27, 28, 29, 30, 32, 33, 34, 35, 36, 38, 39, 40),(1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 14, 15, 17, 18, 19, 20, 21, 22, 23, 24, 25, 26, 27, 28, 29, 30, 31, 32, 34, 35, 36, 39, 40),(3, 4, 5, 6, 7, 8, 9, 11, 12, 13, 14, 16, 17, 18, 19, 21, 22, 23, 24, 25, 26, 27, 28, 29, 30, 31, 32, 34, 35, 36, 37, 38, 40),(0, 1, 2, 6, 7, 8, 9, 10, 11, 13, 14, 16, 17, 18, 19, 21, 22, 24, 25, 26, 27, 28, 31, 32, 37, 38),(0, 2, 3, 4, 5, 6, 9, 10, 14, 15, 17, 19, 21, 22, 27, 28, 29, 32, 38, 39, 40),(1, 3, 4, 5, 6, 7, 10, 12, 14, 15, 17, 19, 20, 21, 22, 23, 24, 25, 26, 28, 29, 30, 31, 32, 36, 37, 39),(6, 7, 8, 9, 10, 13, 14, 15, 17, 18, 19, 21, 22, 24, 26, 27, 28, 31, 32, 33, 35, 38, 39, 40)]
    # 20 iterations
#    bestSets = [(1, 2, 4, 6, 11, 12, 18, 19, 21, 22, 24, 26, 27, 29, 31, 32, 33, 34, 40),(0, 1, 2, 4, 8, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25, 26, 27, 28, 30, 32, 33, 34, 35, 36, 37, 38, 40),(3, 4, 5, 10, 11, 12, 13, 14, 15, 17, 19, 20, 21, 24, 27, 28, 31, 32, 33, 36, 40),(2, 3, 4, 5, 6, 11, 13, 15, 16, 19, 20, 23, 24, 25, 27, 32, 38, 39, 40),(0, 4, 6, 7, 8, 9, 13, 14, 16, 19, 21, 22, 25, 26, 27, 28, 33, 35),(8, 13, 14, 16, 18, 22, 23, 24, 25, 26, 27, 28, 29, 30, 31, 32, 33, 34, 35, 36, 37, 38, 39, 40),(0, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 14, 16, 17, 18, 19, 20, 21, 22, 23, 24, 27, 28, 30, 32, 34, 35, 38, 39, 40),(1, 4, 6, 7, 8, 11, 13, 14, 18, 19, 21, 22, 23, 24, 25, 26, 27, 28, 31, 35, 36, 37, 38),(0, 7, 8, 9, 12, 14, 17, 18, 19, 21, 22, 23, 24, 26, 27, 28, 29, 31, 36, 38),(0, 1, 4, 6, 7, 9, 11, 13, 17, 19, 21, 22, 24, 26, 27, 28, 29, 31, 32, 33, 34, 35, 36, 37),(0, 1, 3, 4, 5, 9, 12, 13, 14, 16, 18, 19, 24, 26, 27, 28, 29, 32, 35, 36, 39, 40),(0, 1, 2, 3, 4, 5, 8, 9, 10, 11, 12, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25, 27, 28, 30, 32, 33, 35, 36, 37, 39),(32, 19, 4, 38, 27, 31, 13, 15),(1, 3, 4, 5, 6, 7, 8, 9, 11, 12, 15, 16, 19, 20, 21, 22, 23, 26, 27, 28, 31, 33, 34, 35, 36),(1, 3, 4, 7, 8, 10, 11, 12, 13, 14, 18, 19, 21, 22, 23, 24, 25, 27, 28, 29, 30, 31, 32, 33, 34, 35, 36, 37, 38, 39, 40),(0, 2, 3, 4, 5, 6, 8, 11, 12, 13, 14, 15, 16, 18, 19, 20, 21, 22, 24, 25, 27, 28, 30, 31, 35, 36, 39, 40),(3, 4, 5, 6, 7, 8, 9, 11, 15, 18, 19, 20, 21, 22, 24, 26, 27, 32, 33, 35),(0, 2, 6, 8, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25, 26, 27, 28, 29, 30, 31, 32, 33, 34, 35, 36, 38, 39, 40),(0, 1, 2, 4, 6, 7, 8, 9, 11, 12, 13, 14, 15, 16, 18, 19, 20, 21, 22, 23, 24, 26, 27, 28, 29, 30, 31, 32, 33, 34, 35, 36, 37, 38, 39, 40),(0, 2, 4, 6, 9, 10, 11, 12, 14, 15, 16, 17, 19, 20, 21, 23, 25, 26, 27, 28, 29, 32, 33, 34, 35, 37, 39, 40)]
    # 50 iterations
#    bestSets = [(0, 2, 3, 4, 5, 8, 11, 15, 16, 18, 19, 20, 22, 27, 28, 29, 32, 36, 38),(0, 1, 2, 3, 4, 5, 8, 11, 12, 15, 17, 18, 19, 20, 21, 23, 24, 27, 28, 30, 32, 33, 36, 37, 38, 39, 40),(0, 1, 7, 9, 11, 12, 13, 14, 15, 16, 18, 19, 20, 21, 22, 23, 24, 25, 26, 27, 28, 29, 30, 31, 33, 34, 35, 36, 37, 38, 39, 40),(8, 10, 11, 12, 16, 17, 18, 19, 20, 21, 22, 24, 25, 27, 29, 30, 34, 35, 38),(0, 2, 3, 4, 5, 13, 14, 15, 18, 19, 20, 22, 24, 26, 27, 28, 33, 39),(0, 4, 6, 7, 8, 9, 11, 13, 14, 16, 18, 19, 21, 22, 23, 24, 26, 27, 28, 31),(0, 6, 7, 9, 11, 12, 14, 16, 19, 22, 24, 26, 27, 28, 29, 30, 31, 34, 36, 37, 38),(0, 2, 3, 4, 5, 9, 10, 11, 12, 14, 17, 18, 19, 20, 22, 23, 24, 25, 27, 28, 29, 30, 35, 37, 38, 39),(0, 1, 3, 4, 5, 11, 12, 14, 15, 17, 19, 20, 21, 22, 23, 24, 27, 28, 29, 32, 33, 34, 36, 37),(0, 1, 2, 3, 4, 5, 7, 8, 12, 14, 16, 17, 19, 22, 23, 26, 27, 28, 29, 32, 35, 39, 40),(0, 1, 3, 4, 7, 8, 11, 14, 15, 16, 19, 22, 23, 24, 26, 27, 28, 31, 40),(1, 3, 4, 6, 14, 17, 19, 21, 22, 23, 24, 26, 27, 28, 31, 35, 38, 40),(0, 4, 7, 8, 11, 13, 14, 16, 19, 21, 22, 23, 24, 27, 28, 38, 40),(0, 2, 3, 4, 5, 8, 11, 15, 16, 18, 19, 20, 22, 27, 28, 29, 32, 36, 38),(0, 1, 2, 3, 4, 5, 8, 11, 12, 15, 17, 18, 19, 20, 21, 23, 24, 27, 28, 30, 32, 33, 36, 37, 38, 39, 40),(0, 2, 8, 9, 11, 13, 17, 18, 19, 21, 22, 24, 27, 31, 32, 33, 39),(33, 35, 40, 9, 11, 15, 17, 18, 19, 20, 21, 22, 23, 24, 27, 28),(2, 3, 4, 5, 6, 11, 15, 16, 19, 20, 24, 26, 27, 28, 29, 30, 31, 32, 33, 34, 35, 38, 39, 40),(0, 1, 2, 4, 6, 7, 8, 9, 11, 12, 13, 14, 16, 17, 19, 21, 22, 23, 24, 25, 27, 29, 30, 31, 32, 34, 35, 36, 38, 40),(1, 2, 4, 13, 14, 17, 19, 20, 22, 24, 26, 27, 28, 30, 31, 34, 35, 36, 38),(2, 6, 7, 8, 9, 10, 12, 13, 14, 15, 16, 17, 19, 20, 21, 22, 23, 24, 25, 26, 27, 28, 29, 30, 31, 32, 33, 35, 36, 37, 38),(0, 6, 9, 13, 14, 15, 17, 19, 21, 22, 23, 26, 27, 28, 33, 38, 40),(32, 1, 2, 10, 12, 14, 19, 20, 21, 22, 23, 24, 27, 28),(32, 17, 19, 38, 27, 28, 12),(2, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 18, 19, 21, 22, 23, 24, 26, 27, 28, 29, 30, 32, 33, 34, 35, 36, 37, 38, 39),(0, 3, 4, 5, 8, 11, 12, 14, 17, 18, 19, 20, 21, 22, 23, 24, 25, 26, 27, 28, 30, 31, 34, 35, 36, 38, 39, 40),(0, 2, 38, 7, 6, 9, 8, 15, 19, 22, 25, 26, 31),(6, 7, 9, 11, 12, 17, 18, 19, 21, 22, 24, 27, 31),(3, 4, 5, 6, 7, 9, 14, 17, 19, 21, 22, 23, 24, 27, 28, 29, 35, 36, 38),(0, 1, 3, 4, 5, 6, 7, 8, 9, 12, 13, 14, 15, 18, 19, 20, 21, 22, 23, 24, 26, 27, 28, 33, 35, 37, 38, 39, 40),(0, 1, 4, 7, 14, 15, 16, 17, 19, 20, 21, 23, 24, 25, 26, 27, 28, 30, 32, 33, 34, 37, 40),(0, 1, 2, 3, 4, 5, 8, 11, 12, 13, 14, 16, 17, 18, 19, 20, 22, 23, 25, 26, 27, 28, 29, 30, 31, 32, 34, 35, 36, 38, 40),(0, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 16, 17, 18, 19, 20, 21, 22, 23, 24, 26, 27, 29, 30, 31, 35, 36, 38),(1, 2, 35, 4, 37, 38, 8, 12, 14, 19, 22, 23, 27, 28, 29),(0, 7, 8, 13, 14, 16, 17, 18, 19, 20, 21, 22, 23, 24, 25, 26, 27, 28, 31, 32),(38, 19, 22, 23, 24, 25, 26, 27, 29),(0, 36, 4, 38, 8, 40, 9, 12, 13, 16, 19, 22, 27),(0, 2, 4, 5, 8, 9, 10, 11, 13, 14, 16, 18, 19, 20, 22, 23, 24, 27, 28, 29, 32, 38, 40),(0, 1, 3, 4, 5, 6, 7, 11, 14, 15, 16, 17, 19, 22, 24, 26, 27, 28, 29, 30, 31, 33, 35, 38),(0, 2, 3, 4, 5, 6, 8, 11, 12, 15, 17, 18, 19, 20, 21, 22, 23, 26, 27, 29, 30, 31, 32, 34, 36, 38, 39),(0, 2, 3, 4, 7, 11, 16, 18, 19, 21, 22, 24, 26, 27, 32, 33, 35, 40),(0, 3, 4, 5, 8, 9, 13, 16, 19, 20, 22, 23, 26, 27, 31, 32, 38, 39, 40),(0, 2, 6, 7, 9, 10, 11, 13, 15, 16, 17, 19, 21, 22, 26, 27, 28, 29, 31, 32, 34, 36),(2, 19, 38, 23, 24, 29),(1, 3, 4, 5, 6, 7, 8, 11, 14, 16, 18, 19, 21, 22, 23, 24, 27, 28, 30, 31, 32, 34, 35, 36, 38),(0, 1, 3, 4, 5, 7, 9, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20, 23, 24, 27, 28, 29, 30, 31, 32, 33, 34, 35, 36, 38),(32, 0, 2, 6, 19, 22, 23, 27, 29),(32, 1, 2, 4, 38, 39, 19, 23, 24, 27, 28),(0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 11, 12, 15, 16, 19, 21, 23, 24, 27, 29, 30, 31, 32, 33, 34, 35, 36),(1, 2, 4, 6, 8, 11, 12, 13, 14, 15, 16, 17, 18, 19, 21, 22, 23, 24, 26, 27, 28, 29, 30, 31, 34, 35, 40)]
    # read combinasion file file
#   data = pd.read_excel('r.xlsx',header=0,sheetname='10_iterations')
#   bestSets=tuple(data.iloc[0:,0])
    #**************************************************************************

#    bestSets = pd.read_excel('r.xlsx',header=0,sheetname='10_iterations')
    # find unique values of all bestSets and get maximum set size
    unique = set()
    maxSetSize = 0
    for i in bestSets:
        unique.update(set(i))
        if len(i) > maxSetSize:
            maxSetSize = len(i)
    print('Number of unique elements found = {}'.format(len(unique)))

    # create a list of all possible combinations with unique values
    hist    = pd.DataFrame()
    for j in list(range(1,maxSetSize+1)):
        print('Searching for all subsets of '+str(j)+' elements...')
        subset = list(itertools.combinations(unique,j))
        print('    {} subsets found'.format(len(subset)))
        print('Computing histogram...')
        df=[] # list to contain histogram of subsets of size j and occurences
        for k in subset:
            df.append([k,0])
            for l in bestSets:
                if set(k).issubset(l):
                    df[-1][1] += 1
            if df[-1][1] == 0:
                del df[-1]
        df1 = pd.DataFrame(data=df,
                           columns=['Subset_'+str(j),'Occurence_'+str(j)])
#        hist = pd.concat([hist,df1],axis=1)

        print('Saving histogram to excel...')
        df1.to_excel('subset_'+str(j)+'.xlsx')

        from slpClass_toolbox import PlotHist
        PlotHist(df1[1],df1[0],'Subsets occurences','Comb_Hist_'+str(j)+'.png')
#        hist.to_excel('subset_'+str(j)+'.xlsx')
        clr() # remove when finish testing

    # used only if there was enough memory to keep all possible subsets
#    print('Saving histogram to excel...')
#    hist.to_excel('subset_hist.xlsx')

    return hist